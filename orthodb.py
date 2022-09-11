import os.path as osp
from os import path, makedirs
import sys
from dataclasses import dataclass
from typing import List

from fasta import read_records, write_records, Record
from taxa import get_lineage_info
from utilities import remote_call, safe_phylo_read, _self_path, wait
from openpyxl import load_workbook

base = 'https://v101.orthodb.org/'


@dataclass
class GeneAnnotation:
    pub_og_id: str
    og_name: str
    level_txid: str
    organism_txid: str
    organism_name: str
    int_prot_id: str
    pub_gene_id: str
    description: str
    sequence: str


async def search(query: str,
                 ncbi: int = 0,
                 level: int = None,
                 skip: int = None,
                 limit: int = 1000,
                 universal: float = None,
                 singlecopy: float = None) -> List[str]:
    """
    /search
    Arguments
        query
            full query string
        ncbi
            flag - if 0, then generic search, if 1 the query is assumed to be a NCBI gene id
        level
            NCBI taxon id of the clade
        skip
            number of hits to skip
        limit
            maximum nr of hits (cluster ids) to return - default is 1000
        universal
            phyloprofile filter, present in 1.0, 0.9, 0.8 of all species in the clade
        singlecopy
            phyloprofile filter, singlecopy in 1.0, 0.9, 0.8 of all species in the clade
    Returns
        a list of clusters, the maximum number of clusters is defined by 'limit'
    Description
        This finds all cluster id's matching a given query.
    """

    res = await remote_call(base + 'search',
                            True,
                            query=query,
                            ncbi=ncbi,
                            level=level,
                            skip=skip,
                            limit=limit,
                            universal=universal,
                            singlecopy=singlecopy)

    return res['data']


async def annotations(id: str,
                      species: List[str] = None,
                      skip: int = None,
                      limit: int = None,
                      universal: float = None,
                      singlecopy: float = None) -> List[GeneAnnotation]:
    """
    /tab
    Arguments
        id
            OrthoDB cluster id
        species
            list of NCBI species taxonomy id's
        skip
            number of hits to skip
        limit
            maximum nr of hits (cluster ids) to return - default is 1000
        universal
            phyloprofile filter, present in 1.0, 0.9, 0.8 of all species in the clade
        singlecopy
            phyloprofile filter, singlecopy in 1.0, 0.9, 0.8 of all species in the clade

    Returns
        tab-separated table of gene annotations
    """

    res = await remote_call(base + 'tab',
                            False,
                            id=id,
                            long=1,
                            species=species if species is None else ",".join(species),
                            skip=skip,
                            limit=limit,
                            universal=universal,
                            singlecopy=singlecopy)

    data = []
    first = True
    for line in res.splitlines():
        if first:
            first = False
            continue

        split = line.split('\t')
        data.append(GeneAnnotation(*split))

    return data


async def seqs_remote(odb_id: str, translate: bool = True) -> List[Record]:

    def get_name(annotation, t):
        if t:
            return get_lineage_info(int(annotation.organism_txid.split('_')[0]))['species'].upper().replace(' ', '_')
        else:
            return f"{annotation.int_prot_id} | {annotation.organism_name} | {annotation.description} | {annotation.pub_gene_id} | {annotation.og_name}"

    annotation_records = await annotations(odb_id, species=[odb_id.split("at")[1]])
    records = [Record(get_name(a, translate), a.sequence) for a in annotation_records]
    return records


# Custom routines that are specific to mammalian ERC analysis
async def get_all_mammalia_orthogroups(universal=1.0, singlecopy=1.0) -> List[str]:
    # Recommend to make universal and singlecopy each 0.8 so that the orthogroups are present in 80% of mammals and are
    # singlecopy in 80% of mammals
    mammalia = 40674

    return await search('', level=mammalia, limit=10000000, universal=universal, singlecopy=singlecopy)


async def odb_seqs(id, remove_multicopy=False, as_filepath=False, force_download=False):
    if not as_filepath:
        if force_download:
            return await seqs_remote(id, False)
        try:
            recs = read_records(osp.join('/scratch/avarela/mammals_pairwise_erc/lib/nuc', id + '.fa'))
            assert len(recs) > 0
            return recs
        except:
            try:
                recs = read_records(osp.join('/scratch/avarela/mammals_pairwise_erc/lib/mito', id + '.fa'))
                assert len(recs) > 0
                return recs
            except:
                assert not osp.exists(osp.join('/scratch/avarela/mammals_pairwise_erc/lib/mito', id + '.fa'))
                downloaded = await seqs_remote(id)
                return downloaded

    renamed = list()

    if not osp.exists(id):
        return None

    was_txid = None
    for record in read_records(id):
        if was_txid is None:
            try:
                i = int(record.title.split("_")[0])
                was_txid = True
            except:
                was_txid = False
                break
        renamed.append(Record(get_lineage_info(int(record.title.split('_')[0]))['species'].upper().replace(' ', '_'), record.sequence))

    if not was_txid:
        assert as_filepath
        return read_records(id)

    if remove_multicopy:
        species = set()
        duped = set()
        for record in renamed:
            if record.title not in species:
                species.add(record.title)
            else:
                duped.add(record.title)
        return [r for r in renamed if r.title not in duped]
    else:
        return renamed


async def write_odb_records(odb: str, output: str, odb_as_filepath: bool = False, translate: bool = True):
    """
    Download records from OrthoDB.

    :param odb: The orthodb id.
    :param output: The output file.
    :param odb_as_filepath: Whether to consider odb as a filepath to a raw orthodb fasta file.
    :param translate: Whether to rename sequences to taxa.
    """
    recs = await odb_seqs(odb, as_filepath=odb_as_filepath)

    if recs is None:
        return

    if translate:
        mammals = set()
        for l in safe_phylo_read(osp.join(_self_path(), "data", "finished_mam_timetree.nwk")).iter_leaf_names():
            mammals.add(l.upper().replace(' ', '_'))

        recs = [r for r in recs if r.title.upper().replace(' ', '_') in mammals]

    write_records(output, recs)
    
def find_cols(sheet, row=1) -> dict:
    cols = {}
    idx = 1
    # go through
    for col in range(1, sheet.max_column + 1):
        cols[sheet.cell(column=col, row=1).value.strip('" ')] = idx
        idx += 1
    return cols
    
def find_paralogs(path:str, sheet_name:str) -> dict:
    """
    Finds all the paralogs from the taxastats excel file. 

    Params:
        path (str): path to taxastats workbook containing list of paralags
        sheet_name (str): _description_
        
    Returns:
        (dict): a dictionary containing the different paralog types (doubles, triples, etc) and their associated proteinids. A list of protein ids per paralog type
    """
    
    paralogs = {
        
    }
    
    wb = load_workbook(path)
    sheet = wb[sheet_name]
    
    cols = find_cols(sheet, 0)
    
    col_protein_id = cols["Protein_ID"]
    col_potential_paralogs = cols["potential_num_paralogs"]
    
    # skip first row (headers)
    for row in range(2, sheet.max_row + 1):
        protein_id = sheet.cell(row=row, column=col_protein_id).value.strip('" ').split()[0]
        paralog_count = int(sheet.cell(row=row, column=col_potential_paralogs).value)
        
        if paralog_count < 2: 
            continue
        key = f"paralog_{paralog_count}'s"
        
        if not key in paralogs:
            paralogs[key] = []
        paralogs[key].append(protein_id)
        
    return paralogs

    
async def get_paralogs(dir, paralogs) -> None:
    for (key, odb_ids) in paralogs.items():
        dirname = key
        dir_path = path.join(dir, dirname)
        for odb_id in odb_ids:
            filename = path.join(dir_path, f"{odb_id}.fa")
            await write_odb_records(odb_id, filename, translate=False)
            
async def get_files(dir_path, fileids) -> None:
    for fileid in fileids: # in this case we using orthoids
        filename = path.join(dir_path, f"{fileid}.fa")
        if not osp.exists(dir_path):
            makedirs(dir_path)
        await write_odb_records(fileid, filename, translate=False)
    


if __name__ == "__main__":
    args = sys.argv[1:]

    if len(args) != 2 and len(args) != 3:
        print("ERROR: Requires 2 or 3 arguments.")
        
        print("For 2 arguments use the following command:")
        print("To download sequences: python3 orthodb.py download ORTHODB_ID")
        print("To rename sequences: python3 orthodb.py rename path/to/fasta.fa")
        
        print("---------------")
        
        print("For 3 arguments use the following command:")
        print("Use command in format: python orthodb_raw.py [excel_file] [excel_sheet] [directory]")
        print("Excel file: Excel file containing the protein ids that are deep paralogs")
        print("Excel sheet: sheet containing the protein ids that are deep paralogs")
        print("Dir: directory to store files downloaded")
            
        
    if len(args) == 2:
        action = args[0]
        arg = args[1]

        if action.lower() == "download":
            output = f"{arg}.fa"
            print("Downloading...")
            wait(write_odb_records(arg, output, translate=False))
            print(f"Saved OrthoDB Entry {arg} to {output}")

        elif action.lower() == "rename":
            if not osp.exists(arg):
                print(f"ERROR: Cannot find sequence file {arg}")
            else:
                split = osp.splitext(arg)
                output = split[0] + ".renamed" + split[1]
                wait(write_odb_records(arg, output, True))
                print(f"Saved renamed fasta to {output}")

        else:
            print(f"ERROR: Cannot recognize action '{action}'")
    else:
        file_path = args[0]
        sheet_name = args[1]
        dir = args[2]
        if not osp.exists(file_path):
            print(f"ERROR: Cannot find sequence file {file_path}")
        if not osp.exists(dir):
            print(f"ERROR: Cannot find directory {dir}")
        
        print("Finding all paralogs")
        paralogs = find_paralogs(file_path, sheet_name)
        print(paralogs)
        # paralogs = {
        #     "paralog_2's": ["177002at40674"]
        # }
        print("Downloading...")
        wait(get_paralogs(dir, paralogs))
